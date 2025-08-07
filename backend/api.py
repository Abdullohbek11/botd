import json
from fastapi import APIRouter, HTTPException
from typing import List, Dict
import os
import requests
from dotenv import load_dotenv
import datetime
from apscheduler.schedulers.background import BackgroundScheduler
import pandas as pd
from io import BytesIO
import base64
import openpyxl

load_dotenv()
BOT_TOKEN = os.getenv("BOT_TOKEN")
GROUP_CHAT_ID_STR = os.getenv("GROUP_CHAT_ID")

# Xavfsiz int ga o'tkazish
if GROUP_CHAT_ID_STR:
    GROUP_CHAT_ID = int(GROUP_CHAT_ID_STR)
else:
    print("ERROR: GROUP_CHAT_ID not found in .env file")
    GROUP_CHAT_ID = None

DATA_DIR = os.path.join(os.path.dirname(__file__), "data")
PRODUCTS_FILE = os.path.join(DATA_DIR, "products.json")
CATEGORIES_FILE = os.path.join(DATA_DIR, "categories.json")
USERS_FILE = os.path.join(DATA_DIR, "users.json")
ORDERS_FILE = os.path.join(DATA_DIR, "orders.json")

router = APIRouter()

def read_json(file_path):
    if not os.path.exists(file_path):
        return []
    with open(file_path, "r", encoding="utf-8") as f:
        return json.load(f)

def write_json(file_path, data):
    with open(file_path, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)

def create_excel_order(order, order_number):
    """Excel fayl yaratish"""
    # Excel uchun ma'lumotlarni tayyorlash
    data = []
    for i, item in enumerate(order.get('items', []), 1):
        name = item.get('name', '')
        quantity = item.get('quantity', 1)
        price = item.get('price') or item.get('product', {}).get('price') or 0
        total = int(price) * int(quantity)
        
        data.append({
            '№': i,
            'MAXSULOT NOMI': name,
            'O\'LCHAM': 'DONA',
            'SONI': quantity,
            'NARXI': price,
            'UMUMIY SUMMA': total
        })
    
    # DataFrame yaratish
    df = pd.DataFrame(data)
    
    # Excel fayl yaratish
    output = BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name='Buyurtma', index=False)
        
        # Excel faylni formatlash
        workbook = writer.book
        worksheet = writer.sheets['Buyurtma']
        
        # Sarlavhalarni formatlash
        for col in range(1, 7):
            cell = worksheet.cell(row=1, column=col)
            cell.font = openpyxl.styles.Font(bold=True)
            cell.fill = openpyxl.styles.PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
        
        # Ustunlar kengligini o'rnatish
        worksheet.column_dimensions['A'].width = 8   # №
        worksheet.column_dimensions['B'].width = 30  # MAXSULOT NOMI
        worksheet.column_dimensions['C'].width = 12  # O'LCHAM
        worksheet.column_dimensions['D'].width = 10  # SONI
        worksheet.column_dimensions['E'].width = 15  # NARXI
        worksheet.column_dimensions['F'].width = 20  # UMUMIY SUMMA
        
        # Jami qatorini qo'shish
        total_row = len(data) + 2
        worksheet.cell(row=total_row, column=1, value="JAMI")
        worksheet.cell(row=total_row, column=4, value=sum(item['SONI'] for item in data))
        worksheet.cell(row=total_row, column=5, value=sum(item['NARXI'] for item in data))
        worksheet.cell(row=total_row, column=6, value=sum(item['UMUMIY SUMMA'] for item in data))
        
        # Jami qatorini formatlash
        for col in range(1, 7):
            cell = worksheet.cell(row=total_row, column=col)
            cell.font = openpyxl.styles.Font(bold=True)
            cell.fill = openpyxl.styles.PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
        
        # Ma'lumotlar qismini qo'shish
        info_row = total_row + 3
        worksheet.cell(row=info_row, column=1, value="YUBORUVCHI:")
        worksheet.cell(row=info_row, column=2, value="O'TKIRBEK")
        worksheet.cell(row=info_row, column=4, value="YUBORUVCHI NOMERI:")
        worksheet.cell(row=info_row, column=5, value="+998979960020")
        
        info_row += 1
        worksheet.cell(row=info_row, column=1, value="QABUL QILUVCHI:")
        worksheet.cell(row=info_row, column=2, value=order.get('customerInfo', {}).get('name', '-'))
        worksheet.cell(row=info_row, column=4, value="Tel raqam:")
        worksheet.cell(row=info_row, column=5, value=order.get('customerInfo', {}).get('phone', '-'))
        
        info_row += 1
        worksheet.cell(row=info_row, column=1, value="Telegram:")
        worksheet.cell(row=info_row, column=2, value="@Bronavia0020")
        worksheet.cell(row=info_row, column=4, value="Jami summa:")
        worksheet.cell(row=info_row, column=5, value=f"{order.get('total', 0)} so'm")
        
        # Ma'lumotlar qismini formatlash
        for row in range(total_row + 3, info_row + 1):
            for col in range(1, 6):
                cell = worksheet.cell(row=row, column=col)
                if col in [1, 4]:  # Label ustunlari
                    cell.font = openpyxl.styles.Font(bold=True)
                    cell.fill = openpyxl.styles.PatternFill(start_color="E6E6E6", end_color="E6E6E6", fill_type="solid")
    
    output.seek(0)
    return output

def send_order_to_group(order, order_number):
    print(f"DEBUG: send_order_to_group chaqirildi")
    print(f"DEBUG: GROUP_CHAT_ID = {GROUP_CHAT_ID}")
    print(f"DEBUG: BOT_TOKEN = {BOT_TOKEN[:10]}...")
    
    if GROUP_CHAT_ID is None:
        print("ERROR: GROUP_CHAT_ID is not set")
        return
    
    if not BOT_TOKEN:
        print("ERROR: BOT_TOKEN is not set")
        return
    
    # Lokatsiya yuborish
    loc = order.get('customerInfo', {}).get('location', '')
    try:
        if loc and ',' in loc:
            lat, lon = map(float, loc.split(','))
            url = f"https://api.telegram.org/bot{BOT_TOKEN}/sendLocation"
            payload = {
                "chat_id": GROUP_CHAT_ID,
                "latitude": lat,
                "longitude": lon
            }
            print(f"DEBUG: Location yuborish - URL: {url}")
            print(f"DEBUG: Location payload: {payload}")
            r = requests.post(url, json=payload, timeout=5)
            print("Telegram location javobi:", r.text)
    except Exception as e:
        print("Telegram location xatolik:", e)
    
    # Xabar yuborish
    text = (
        f"🛒 #{order_number}-chi buyurtma!\n"
        f"<b>Ism:</b> {order.get('customerInfo', {}).get('name', '-')}\n"
        f"<b>Telefon:</b> {order.get('customerInfo', {}).get('phone', '-') }\n"
        f"<b>Manzil:</b> {order.get('customerInfo', {}).get('address', '-') }\n"
        f"<b>Joylashuv:</b> {order.get('customerInfo', {}).get('location', '-')}\n"
        f"<b>Mahsulotlar:</b>\n"
    )
    for item in order.get('items', []):
        name = item.get('name', '')
        quantity = item.get('quantity', 1)
        price = item.get('price') or item.get('product', {}).get('price') or 0
        subtotal = int(price) * int(quantity)
        text += f"- {name} {quantity} dona = {subtotal} so'm\n"
    text += f"\nJami: {order.get('total', 0)} so'm"
    
    url = f"https://api.telegram.org/bot{BOT_TOKEN}/sendMessage"
    payload = {
        "chat_id": GROUP_CHAT_ID,
        "text": text,
        "parse_mode": "HTML"
    }
    print(f"DEBUG: Xabar yuborish - URL: {url}")
    print(f"DEBUG: Xabar payload: {payload}")
    try:
        r = requests.post(url, json=payload, timeout=5)
        print("Telegram API javobi:", r.text)
        if r.status_code != 200:
            print(f"ERROR: Telegram API xatolik - status: {r.status_code}")
    except Exception as e:
        print("Telegram API xatolik:", e)
    
    # Excel fayl yaratish va yuborish
    try:
        excel_file = create_excel_order(order, order_number)
        excel_file.seek(0)
        
        # Excel faylni yuborish
        files = {'document': ('buyurtma.xlsx', excel_file, 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')}
        data = {
            'chat_id': GROUP_CHAT_ID,
            'caption': f'📋 #{order_number}-chi buyurtma Excel fayli'
        }
        
        url = f"https://api.telegram.org/bot{BOT_TOKEN}/sendDocument"
        r = requests.post(url, data=data, files=files, timeout=10)
        print("Excel fayl yuborildi:", r.text)
        
    except Exception as e:
        print("Excel fayl yuborishda xatolik:", e)

def get_time_period_stats(start_time, end_time, period_name):
    """Berilgan vaqt oralig'idagi statistikalarni hisoblaydi"""
    orders = read_json(ORDERS_FILE)
    product_stats = {}
    total_sum = 0
    total_orders = 0
    customer_orders = []
    
    for order in orders:
        created = order.get('created_at')
        if not created:
            continue
        try:
            created_dt = datetime.datetime.fromisoformat(created)
        except Exception:
            continue
        if not (start_time <= created_dt < end_time):
            continue
        
        total_orders += 1
        order_sum = 0
        customer_name = order.get('customerInfo', {}).get('name', 'Noma\'lum')
        order_id = order.get('id', 'N/A')
        
        for item in order.get('items', []):
            name = item.get('name', '')
            quantity = int(item.get('quantity', 1))
            price = int(item.get('price', 0))
            item_total = quantity * price
            order_sum += item_total
            
            if name not in product_stats:
                product_stats[name] = {'quantity': 0, 'total': 0}
            product_stats[name]['quantity'] += quantity
            product_stats[name]['total'] += item_total
        
        total_sum += order_sum
        
        # Mijoz buyurtmasini qo'shamiz
        items_str = ', '.join([f"{item.get('name', '')} — {item.get('quantity', 1)} dona" for item in order.get('items', [])])
        customer_orders.append(f"{customer_name} (#{order_id}): {items_str}")
    
    return product_stats, total_sum, total_orders, customer_orders

def send_morning_stats():
    """Ertalab 7:00-10:00 oralig'idagi statistika"""
    try:
        now = datetime.datetime.now()
        today_7 = now.replace(hour=7, minute=0, second=0, microsecond=0)
        today_10 = now.replace(hour=10, minute=0, second=0, microsecond=0)
        
        product_stats, total_sum, total_orders, customer_orders = get_time_period_stats(today_7, today_10, "Ertalab")
        
        if not product_stats:
            stats_text = 'Ertalab 7:00-10:00 oralig\'ida buyurtmalar yo\'q.'
        else:
            stats_text = f'📊 Ertalab 7:00-10:00 oralig\'idagi buyurtmalar:\n\n'
            
            stats_text += 'Mahsulotlar:\n'
            for name, data in product_stats.items():
                stats_text += f'{name} — {data["quantity"]} dona, {data["total"]} so\'m\n'
            
            stats_text += f'\nJami buyurtmalar: {total_orders} ta\n'
            stats_text += f'Jami summa: {total_sum} so\'m\n'
            
            if customer_orders:
                stats_text += '\nMijozlar:\n'
                for customer_order in customer_orders:
                    stats_text += f'- {customer_order}\n'
        
        url = f"https://api.telegram.org/bot{BOT_TOKEN}/sendMessage"
        payload = {
            "chat_id": GROUP_CHAT_ID,
            "text": stats_text
        }
        r = requests.post(url, json=payload, timeout=5)
        print("Ertalab statistikasi yuborildi:", r.text)
        
    except Exception as e:
        print("Ertalab statistika xatolik:", e)

def send_mid_morning_stats():
    """Kun o\'rtasi 10:00-13:00 oralig'idagi statistika"""
    try:
        now = datetime.datetime.now()
        today_10 = now.replace(hour=10, minute=0, second=0, microsecond=0)
        today_13 = now.replace(hour=13, minute=0, second=0, microsecond=0)
        
        product_stats, total_sum, total_orders, customer_orders = get_time_period_stats(today_10, today_13, "Kun o'rtasi")
        
        if not product_stats:
            stats_text = 'Kun o\'rtasi 10:00-13:00 oralig\'ida buyurtmalar yo\'q.'
        else:
            stats_text = f'📊 Kun o\'rtasi 10:00-13:00 oralig\'idagi buyurtmalar:\n\n'
            
            stats_text += 'Mahsulotlar:\n'
            for name, data in product_stats.items():
                stats_text += f'{name} — {data["quantity"]} dona, {data["total"]} so\'m\n'
            
            stats_text += f'\nJami buyurtmalar: {total_orders} ta\n'
            stats_text += f'Jami summa: {total_sum} so\'m\n'
            
            if customer_orders:
                stats_text += '\nMijozlar:\n'
                for customer_order in customer_orders:
                    stats_text += f'- {customer_order}\n'
        
        url = f"https://api.telegram.org/bot{BOT_TOKEN}/sendMessage"
        payload = {
            "chat_id": GROUP_CHAT_ID,
            "text": stats_text
        }
        r = requests.post(url, json=payload, timeout=5)
        print("Kun o'rtasi statistikasi yuborildi:", r.text)
        
    except Exception as e:
        print("Kun o'rtasi statistika xatolik:", e)

def send_afternoon_stats():
    """Tushdan keyin 13:00-16:00 oralig'idagi statistika"""
    try:
        now = datetime.datetime.now()
        today_13 = now.replace(hour=13, minute=0, second=0, microsecond=0)
        today_16 = now.replace(hour=16, minute=0, second=0, microsecond=0)
        
        product_stats, total_sum, total_orders, customer_orders = get_time_period_stats(today_13, today_16, "Tushdan keyin")
        
        if not product_stats:
            stats_text = 'Tushdan keyin 13:00-16:00 oralig\'ida buyurtmalar yo\'q.'
        else:
            stats_text = f'📊 Tushdan keyin 13:00-16:00 oralig\'idagi buyurtmalar:\n\n'
            
            stats_text += 'Mahsulotlar:\n'
            for name, data in product_stats.items():
                stats_text += f'{name} — {data["quantity"]} dona, {data["total"]} so\'m\n'
            
            stats_text += f'\nJami buyurtmalar: {total_orders} ta\n'
            stats_text += f'Jami summa: {total_sum} so\'m\n'
            
            if customer_orders:
                stats_text += '\nMijozlar:\n'
                for customer_order in customer_orders:
                    stats_text += f'- {customer_order}\n'
        
        url = f"https://api.telegram.org/bot{BOT_TOKEN}/sendMessage"
        payload = {
            "chat_id": GROUP_CHAT_ID,
            "text": stats_text
        }
        r = requests.post(url, json=payload, timeout=5)
        print("Tushdan keyin statistikasi yuborildi:", r.text)
        
    except Exception as e:
        print("Tushdan keyin statistika xatolik:", e)

def send_evening_stats():
    """Kechqurun 16:00-19:00 oralig'idagi statistika"""
    try:
        now = datetime.datetime.now()
        today_16 = now.replace(hour=16, minute=0, second=0, microsecond=0)
        today_19 = now.replace(hour=19, minute=0, second=0, microsecond=0)
        
        product_stats, total_sum, total_orders, customer_orders = get_time_period_stats(today_16, today_19, "Kechqurun")
        
        if not product_stats:
            stats_text = 'Kechqurun 16:00-19:00 oralig\'ida buyurtmalar yo\'q.'
        else:
            stats_text = f'📊 Kechqurun 16:00-19:00 oralig\'idagi buyurtmalar:\n\n'
            
            stats_text += 'Mahsulotlar:\n'
            for name, data in product_stats.items():
                stats_text += f'{name} — {data["quantity"]} dona, {data["total"]} so\'m\n'
            
            stats_text += f'\nJami buyurtmalar: {total_orders} ta\n'
            stats_text += f'Jami summa: {total_sum} so\'m\n'
            
            if customer_orders:
                stats_text += '\nMijozlar:\n'
                for customer_order in customer_orders:
                    stats_text += f'- {customer_order}\n'
        
        url = f"https://api.telegram.org/bot{BOT_TOKEN}/sendMessage"
        payload = {
            "chat_id": GROUP_CHAT_ID,
            "text": stats_text
        }
        r = requests.post(url, json=payload, timeout=5)
        print("Kechqurun statistikasi yuborildi:", r.text)
        
    except Exception as e:
        print("Kechqurun statistika xatolik:", e)

def send_night_stats():
    """Tungi 19:00-7:00 oralig'idagi statistika"""
    try:
        now = datetime.datetime.now()
        today_19 = now.replace(hour=19, minute=0, second=0, microsecond=0)
        tomorrow_7 = (now + datetime.timedelta(days=1)).replace(hour=7, minute=0, second=0, microsecond=0)
        
        product_stats, total_sum, total_orders, customer_orders = get_time_period_stats(today_19, tomorrow_7, "Tungi")
        
        if not product_stats:
            stats_text = 'Tungi 19:00-7:00 oralig\'ida buyurtmalar yo\'q.'
        else:
            stats_text = f'📊 Tungi 19:00-7:00 oralig\'idagi buyurtmalar:\n\n'
            
            stats_text += 'Mahsulotlar:\n'
            for name, data in product_stats.items():
                stats_text += f'{name} — {data["quantity"]} dona, {data["total"]} so\'m\n'
            
            stats_text += f'\nJami buyurtmalar: {total_orders} ta\n'
            stats_text += f'Jami summa: {total_sum} so\'m\n'
            
            if customer_orders:
                stats_text += '\nMijozlar:\n'
                for customer_order in customer_orders:
                    stats_text += f'- {customer_order}\n'
        
        url = f"https://api.telegram.org/bot{BOT_TOKEN}/sendMessage"
        payload = {
            "chat_id": GROUP_CHAT_ID,
            "text": stats_text
        }
        r = requests.post(url, json=payload, timeout=5)
        print("Tungi statistikasi yuborildi:", r.text)
        
    except Exception as e:
        print("Tungi statistika xatolik:", e)

def send_all_weeks_stats():
    import datetime
    orders = read_json(ORDERS_FILE)
    # Buyurtmalarni haftalarga ajratamiz
    weeks = {}
    for order in orders:
        created = order.get('created_at')
        if created:
            try:
                created_dt = datetime.datetime.fromisoformat(created)
            except Exception:
                continue
            year, week_num, _ = created_dt.isocalendar()
            key = f"{year}-yil, {week_num}-hafta"
            if key not in weeks:
                weeks[key] = []
            weeks[key].append(order)
    # Har bir hafta uchun statistika tuzamiz
    all_stats = ""
    for week_key in sorted(weeks.keys()):
        week_orders = weeks[week_key]
        product_stats = {}
        total_sum = 0
        total_orders = 0
        customer_orders = []
        for order in week_orders:
            total_orders += 1
            order_sum = 0
            customer_name = order.get('customerInfo', {}).get('name', 'Noma\'lum')
            order_id = order.get('id', 'N/A')
            for item in order.get('items', []):
                name = item.get('name', '')
                quantity = int(item.get('quantity', 1))
                price = int(item.get('price', 0))
                item_total = quantity * price
                order_sum += item_total
                if name not in product_stats:
                    product_stats[name] = {'quantity': 0, 'total': 0}
                product_stats[name]['quantity'] += quantity
                product_stats[name]['total'] += item_total
            total_sum += order_sum
            items_str = ', '.join([f"{item.get('name', '')} — {item.get('quantity', 1)} dona" for item in order.get('items', [])])
            customer_orders.append(f"{customer_name} (#{order_id}): {items_str}")
        all_stats += f"\n==============================\n"
        all_stats += f"📅 {week_key} statistikasi:\n\n"
        if not product_stats:
            all_stats += "Hafta davomida buyurtmalar yo'q.\n"
        else:
            all_stats += 'Mahsulotlar:\n'
            for name, data in product_stats.items():
                all_stats += f'{name} — {data["quantity"]} dona, {data["total"]} so\'m\n'
            all_stats += f'\nJami buyurtmalar: {total_orders} ta\n'
            all_stats += f'Jami summa: {total_sum} so\'m\n'
            if customer_orders:
                all_stats += '\nMijozlar:\n'
                for customer_order in customer_orders:
                    all_stats += f'- {customer_order}\n'
    # txt faylga yozamiz
    txt_file = os.path.join(DATA_DIR, f"all_weeks_stats.txt")
    with open(txt_file, "w", encoding="utf-8") as f:
        f.write(all_stats)
    # Telegramga fayl yuboramiz
    url = f"https://api.telegram.org/bot{BOT_TOKEN}/sendDocument"
    files = {'document': open(txt_file, 'rb')}
    data = {"chat_id": GROUP_CHAT_ID, "caption": "Barcha haftalar statistikasi"}
    try:
        r = requests.post(url, data=data, files=files, timeout=10)
        print("Barcha haftalar statistikasi fayli yuborildi:", r.text)
    except Exception as e:
        print("Barcha haftalar statistikasi fayli xatolik:", e)
    finally:
        files['document'].close()

# --- Scheduler ---
scheduler = BackgroundScheduler()
# 4 ta vaqt oralig'i uchun statistika
scheduler.add_job(send_morning_stats, 'cron', hour=10, minute=0)  # Har kuni 10:00 da (7:00-10:00)
scheduler.add_job(send_mid_morning_stats, 'cron', hour=13, minute=0)  # Har kuni 13:00 da (10:00-13:00)
scheduler.add_job(send_afternoon_stats, 'cron', hour=16, minute=0)  # Har kuni 16:00 da (13:00-16:00)
scheduler.add_job(send_evening_stats, 'cron', hour=19, minute=0)  # Har kuni 19:00 da (16:00-19:00)
scheduler.add_job(send_night_stats, 'cron', hour=7, minute=0)  # Har kuni 7:00 da (19:00-7:00)
scheduler.add_job(send_all_weeks_stats, 'cron', day_of_week='sun', hour=22, minute=0)  # Yakshanba 22:00 da
scheduler.start()

# --- PRODUCTS ---
@router.get("/products")
def get_products():
    return read_json(PRODUCTS_FILE)

@router.post("/products")
def add_product(product: Dict):
    mapped_product = {
        "id": str(product.get("id")),
        "name": product.get("name"),
        "description": product.get("description", ""),
        "price": product.get("price", 0),
        "original_price": product.get("original_price", 0),
        "discount": product.get("discount", 0),
        "image": product.get("image", ""),
        "category_id": product.get("category_id"),
        "in_stock": product.get("inStock", True),
        "rating": product.get("rating", 0),
        "reviews_count": product.get("reviews_count", 0),
        "tannarxi": product.get("tannarxi", 0)
    }
    products = read_json(PRODUCTS_FILE)
    products.append(mapped_product)
    write_json(PRODUCTS_FILE, products)
    return mapped_product

@router.delete("/products/{product_id}")
def delete_product(product_id: str):
    products = read_json(PRODUCTS_FILE)
    products = [p for p in products if str(p.get("id")) != str(product_id)]
    write_json(PRODUCTS_FILE, products)
    return {"success": True}

# --- CATEGORIES ---
@router.get("/categories")
def get_categories():
    return read_json(CATEGORIES_FILE)

@router.post("/categories")
def add_category(category: Dict):
    mapped_category = {
        "id": str(category.get("id")),
        "name": category.get("name", ""),
        "icon": category.get("icon", ""),
        "image": category.get("image", ""),
        "productCount": category.get("productCount", 0)
    }
    categories = read_json(CATEGORIES_FILE)
    categories.append(mapped_category)
    write_json(CATEGORIES_FILE, categories)
    return mapped_category

@router.delete("/categories/{category_id}")
def delete_category(category_id: str):
    categories = read_json(CATEGORIES_FILE)
    categories = [c for c in categories if str(c.get("id")) != str(category_id)]
    write_json(CATEGORIES_FILE, categories)
    return {"success": True}

# --- USERS ---
@router.get("/users")
def get_users():
    return read_json(USERS_FILE)

@router.post("/users")
def add_user(user: Dict):
    users = read_json(USERS_FILE)
    users.append(user)
    write_json(USERS_FILE, users)
    return user 

# --- ORDERS ---
@router.get("/orders")
def get_orders():
    return read_json(ORDERS_FILE)

@router.post("/orders")
def add_order(order: Dict):
    orders = read_json(ORDERS_FILE)
    order_id = len(orders) + 1
    order['id'] = order_id
    order['created_at'] = datetime.datetime.now().isoformat()
    orders.append(order)
    write_json(ORDERS_FILE, orders)
    send_order_to_group(order, order_id)
    return {"message": f"Buyurtma muvaffaqiyatli yuborildi! #{order_id}-chi buyurtma. Tez orada siz bilan bog'lanamiz."} 